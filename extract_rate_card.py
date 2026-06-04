"""
Extract rate card data from an Excel workbook in the input folder.

Prompts for the source file and sheet mappings, then writes a consolidated
workbook to the processing folder.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import pandas as pd

from project_paths import INPUT_DIR, PROCESSING_DIR

DEFAULT_SHEETS = {
    "rate_card": "Rate Card",
    "accessorial_costs": "Extra Charges",
    "zones": "Corporatezoning",
    "address_zones": "Sonderzoning",
}

OUTPUT_SHEET_NAMES = {
    "rate_card": "Rate Card",
    "accessorial_costs": "Accessorial Costs",
    "zones": "Zones",
    "address_zones": "Address Zones",
}


@dataclass
class PricingRoundingRules:
    price_per_shipment_decimals: int | None
    price_per_100kg_decimals: int | None
    prohibit_hundred_kg_round_up: bool
    raw_rule: str | None = None


@dataclass
class ExtractionResult:
    rate_card: pd.DataFrame
    accessorial_costs: pd.DataFrame
    zones: pd.DataFrame
    address_zones: pd.DataFrame
    miscellaneous_pricing_rule: str | None
    pricing_rounding_rules: PricingRoundingRules
    rate_card_header_row_count: int
    source_file: Path
    output_file: Path


def list_input_files() -> list[Path]:
    if not INPUT_DIR.exists():
        INPUT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(INPUT_DIR.glob("*.xlsx")) + sorted(INPUT_DIR.glob("*.xls"))


def list_workbook_sheets(file_path: Path) -> list[str]:
    excel_file = pd.ExcelFile(file_path)
    return excel_file.sheet_names


def resolve_default_sheet(sheet_names: list[str], default_name: str) -> str | None:
    if default_name in sheet_names:
        return default_name

    lowered = {name.lower(): name for name in sheet_names}
    if default_name.lower() in lowered:
        return lowered[default_name.lower()]

    default_lower = default_name.lower()
    for name in sheet_names:
        if default_lower in name.lower():
            return name

    return None


def prompt_choice(prompt: str, options: list[str], default: str | None = None) -> str:
    print(f"\n{prompt}")
    for index, option in enumerate(options, start=1):
        marker = " (default)" if default and option == default else ""
        print(f"  {index}. {option}{marker}")

    while True:
        raw = input("Enter number or exact name: ").strip()
        if not raw and default:
            return default
        if raw.isdigit():
            choice_index = int(raw) - 1
            if 0 <= choice_index < len(options):
                return options[choice_index]
        if raw in options:
            return raw
        print("Invalid choice. Try again.")


def prompt_sheet_name(
    sheet_names: list[str],
    label: str,
    default_name: str,
) -> str:
    suggested = resolve_default_sheet(sheet_names, default_name)

    print(f"\n{label}")
    print("Available tabs:")
    for index, name in enumerate(sheet_names, start=1):
        marker = "  <-- auto-selected" if suggested and name == suggested else ""
        print(f"  {index}. {name}{marker}")

    if suggested:
        print(
            f'\nAuto-selected: "{suggested}" '
            f'(tab name contains "{default_name}")'
        )
        prompt_text = "Press Enter to confirm, or enter tab number: "
    else:
        print(f'\nNo tab matched "{default_name}".')
        prompt_text = "Enter tab number: "

    while True:
        raw = input(prompt_text).strip()
        if not raw:
            if suggested:
                return suggested
            print("Please enter a tab number.")
            continue
        if raw.isdigit():
            choice_index = int(raw) - 1
            if 0 <= choice_index < len(sheet_names):
                return sheet_names[choice_index]
        print("Invalid choice. Enter a valid tab number.")


def find_lane_zoning_row(df: pd.DataFrame) -> int | None:
    for row_index in range(len(df)):
        for column_index in range(df.shape[1]):
            value = df.iloc[row_index, column_index]
            if pd.isna(value):
                continue
            text = str(value).lower()
            if "lane" in text and "zoning" in text:
                return row_index
    return None


def find_rate_card_header_row(df: pd.DataFrame, lane_row: int) -> int | None:
    for row_index in range(lane_row, min(lane_row + 6, len(df))):
        first_cell = df.iloc[row_index, 0]
        if pd.isna(first_cell):
            continue
        if "origin" in str(first_cell).lower():
            return row_index
    return None


def _header_cell_text(value: object) -> str:
    if pd.isna(value) or str(value).strip() == "":
        return ""
    return str(value).replace("\n", " ").strip()


def _make_unique_bottom_level_names(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique_names: list[str] = []

    for name in names:
        if name not in seen:
            seen[name] = 1
            unique_names.append(name)
            continue

        seen[name] += 1
        unique_names.append(f"{name}_{seen[name]}")

    return unique_names


def build_rate_card_column_index(header_band: pd.DataFrame) -> pd.MultiIndex:
    if len(header_band) < 3:
        raise ValueError("Expected 3 header rows above the Rate Card data.")

    top_row = header_band.iloc[0].copy()
    middle_row = header_band.iloc[1].copy()
    bottom_row = header_band.iloc[2].copy()

    pricing_start = None
    pricing_end = len(middle_row)
    for column_index in range(len(middle_row)):
        value = _header_cell_text(middle_row.iloc[column_index])
        if value and "price per" in value.lower():
            if pricing_start is None:
                pricing_start = column_index
            pricing_end = column_index + 1
        elif pricing_start is not None and value:
            pricing_end = column_index + 1

    if pricing_start is None:
        pricing_start = 13
        pricing_end = len(middle_row)

    lane_section = top_row.iloc[:pricing_start].copy()
    if lane_section.replace("", pd.NA).notna().any():
        lane_section = lane_section.replace("", pd.NA).ffill().bfill().fillna("")
    top_row.iloc[:pricing_start] = lane_section

    pricing_top = top_row.iloc[pricing_start:pricing_end].replace("", pd.NA).ffill().bfill().fillna("")
    top_row.iloc[pricing_start:pricing_end] = pricing_top

    middle_pricing = (
        middle_row.iloc[pricing_start:pricing_end].replace("", pd.NA).ffill().bfill().fillna("")
    )
    middle_row.iloc[pricing_start:pricing_end] = middle_pricing

    bottom_names: list[str] = []
    for column_index in range(len(bottom_row)):
        name = _header_cell_text(bottom_row.iloc[column_index])
        bottom_names.append(name or f"Column_{column_index + 1}")
    bottom_names = _make_unique_bottom_level_names(bottom_names)

    column_tuples: list[tuple[str, str, str]] = []
    for column_index in range(len(bottom_row)):
        column_tuples.append(
            (
                _header_cell_text(top_row.iloc[column_index]),
                _header_cell_text(middle_row.iloc[column_index]),
                bottom_names[column_index],
            )
        )

    return pd.MultiIndex.from_tuples(
        column_tuples,
        names=["Service Block", "Pricing Type", "Column"],
    )


def extract_rate_card_df(
    df_raw: pd.DataFrame,
    file_path: Path | None = None,
    sheet_name: str | None = None,
) -> tuple[pd.DataFrame, int]:
    lane_row = find_lane_zoning_row(df_raw)
    if lane_row is None:
        raise ValueError("Could not find a 'Lane Zoning' marker in the Rate Card tab.")

    header_row = find_rate_card_header_row(df_raw, lane_row)
    if header_row is None:
        header_row = lane_row + 2

    header_row_count = header_row - lane_row + 1
    header_band = df_raw.iloc[lane_row : header_row + 1].copy()

    if file_path is not None and sheet_name is not None:
        header_band = apply_excel_merged_cells(
            header_band,
            file_path,
            sheet_name,
            start_row=lane_row,
        )

    column_index = build_rate_card_column_index(header_band)
    rate_card = df_raw.iloc[header_row + 1 :].copy()
    rate_card.columns = column_index
    rate_card = rate_card.dropna(how="all").reset_index(drop=True)
    return rate_card, header_row_count


def apply_excel_merged_cells(
    df_slice: pd.DataFrame,
    file_path: Path,
    sheet_name: str,
    start_row: int,
) -> pd.DataFrame:
    """Copy merged-cell values into all cells covered by each Excel merge range."""
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    try:
        worksheet = workbook[sheet_name]
        result = df_slice.copy()

        for merge_range in worksheet.merged_cells.ranges:
            value = worksheet.cell(merge_range.min_row, merge_range.min_col).value
            for excel_row in range(merge_range.min_row, merge_range.max_row + 1):
                for excel_col in range(merge_range.min_col, merge_range.max_col + 1):
                    slice_row = excel_row - 1 - start_row
                    slice_col = excel_col - 1
                    if slice_row < 0 or slice_row >= len(result):
                        continue
                    if slice_col < 0 or slice_col >= result.shape[1]:
                        continue
                    result.iloc[slice_row, slice_col] = value

        return result
    finally:
        workbook.close()


def extract_miscellaneous_pricing_rule(df: pd.DataFrame) -> str | None:
    for row_index in range(len(df)):
        label = df.iloc[row_index, 0]
        if pd.isna(label):
            continue
        if not str(label).strip().lower().startswith("miscellaneous"):
            continue

        for column_index in range(1, df.shape[1]):
            value = df.iloc[row_index, column_index]
            if pd.isna(value):
                continue
            text = str(value).strip()
            if "price per shipment" in text.lower():
                return text
    return None


def parse_pricing_rounding_rules(rule: str | None) -> PricingRoundingRules:
    if not rule:
        return PricingRoundingRules(None, None, False, rule)

    rule_lower = rule.lower()
    prohibit_hundred_kg_round_up = (
        "rounding up to full hundred kg is not permitted" in rule_lower
    )

    shipment_decimals = _parse_decimal_places(
        rule,
        r"price\s+per\s+shipment[^:]*:\s*(\d+)\s+decimal\s+places?",
    )
    hundred_kg_decimals = _parse_decimal_places(
        rule,
        r"price\s+per\s+100\s*kg[^:]*:\s*(\d+)\s+decimal\s+places?",
    )

    shared_match = re.search(r"(\d+)\s+decimal\s+places?", rule, re.IGNORECASE)
    shared_decimals = int(shared_match.group(1)) if shared_match else None

    if shipment_decimals is None:
        shipment_decimals = shared_decimals
    if hundred_kg_decimals is None:
        hundred_kg_decimals = shared_decimals

    return PricingRoundingRules(
        price_per_shipment_decimals=shipment_decimals,
        price_per_100kg_decimals=hundred_kg_decimals,
        prohibit_hundred_kg_round_up=prohibit_hundred_kg_round_up,
        raw_rule=rule,
    )


def _parse_decimal_places(rule: str, pattern: str) -> int | None:
    match = re.search(pattern, rule, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


def _round_numeric_value(value: object, decimal_places: int) -> object:
    if pd.isna(value):
        return value

    numeric_value = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric_value):
        return value

    quantizer = Decimal("1").scaleb(-decimal_places)
    rounded = Decimal(str(numeric_value)).quantize(quantizer, rounding=ROUND_HALF_UP)
    return float(rounded)


def _pricing_type_decimals(
    pricing_type: str,
    rules: PricingRoundingRules,
) -> int | None:
    pricing_type_lower = pricing_type.lower()

    if "price per shipment" in pricing_type_lower:
        return rules.price_per_shipment_decimals
    if "price per 100kg" in pricing_type_lower:
        return rules.price_per_100kg_decimals
    if pricing_type_lower == "max":
        return rules.price_per_shipment_decimals

    return None


def apply_pricing_rounding(
    rate_card: pd.DataFrame,
    rules: PricingRoundingRules,
) -> pd.DataFrame:
    if not isinstance(rate_card.columns, pd.MultiIndex):
        return rate_card

    rounded = rate_card.copy()

    for column in rounded.columns:
        pricing_type = str(column[1]) if len(column) > 1 else ""
        decimal_places = _pricing_type_decimals(pricing_type, rules)
        if decimal_places is None:
            continue

        rounded[column] = rounded[column].map(
            lambda value: _round_numeric_value(value, decimal_places)
        )

    return rounded


def read_sheet_as_dataframe(file_path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(file_path, sheet_name=sheet_name, header=0)


def extract_rate_card_workbook(
    file_path: Path,
    sheet_mapping: dict[str, str],
) -> ExtractionResult:
    rate_card_raw = pd.read_excel(
        file_path,
        sheet_name=sheet_mapping["rate_card"],
        header=None,
    )

    rate_card_df, header_row_count = extract_rate_card_df(
        rate_card_raw,
        file_path=file_path,
        sheet_name=sheet_mapping["rate_card"],
    )
    accessorial_costs_df = read_sheet_as_dataframe(
        file_path,
        sheet_mapping["accessorial_costs"],
    )
    zones_df = read_sheet_as_dataframe(file_path, sheet_mapping["zones"])
    address_zones_df = read_sheet_as_dataframe(
        file_path,
        sheet_mapping["address_zones"],
    )
    miscellaneous_pricing_rule = extract_miscellaneous_pricing_rule(rate_card_raw)
    pricing_rounding_rules = parse_pricing_rounding_rules(miscellaneous_pricing_rule)
    rate_card_df = apply_pricing_rounding(rate_card_df, pricing_rounding_rules)

    PROCESSING_DIR.mkdir(parents=True, exist_ok=True)
    output_file = PROCESSING_DIR / f"{file_path.stem}_extracted.xlsx"

    metadata_rows = [
        {"Field": "Source file", "Value": file_path.name},
        {
            "Field": "Miscellaneous pricing rule",
            "Value": miscellaneous_pricing_rule or "",
        },
        {
            "Field": "Price per shipment decimal places",
            "Value": pricing_rounding_rules.price_per_shipment_decimals or "",
        },
        {
            "Field": "Price per 100kg decimal places",
            "Value": pricing_rounding_rules.price_per_100kg_decimals or "",
        },
        {
            "Field": "Hundred kg round-up prohibited",
            "Value": pricing_rounding_rules.prohibit_hundred_kg_round_up,
        },
        {
            "Field": "Rate Card column header levels",
            "Value": header_row_count,
        },
    ]
    metadata_df = pd.DataFrame(metadata_rows)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        rate_card_df.to_excel(
            writer,
            sheet_name=OUTPUT_SHEET_NAMES["rate_card"],
            merge_cells=True,
        )
        accessorial_costs_df.to_excel(
            writer,
            sheet_name=OUTPUT_SHEET_NAMES["accessorial_costs"],
            index=False,
        )
        zones_df.to_excel(writer, sheet_name=OUTPUT_SHEET_NAMES["zones"], index=False)
        address_zones_df.to_excel(
            writer,
            sheet_name=OUTPUT_SHEET_NAMES["address_zones"],
            index=False,
        )
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False)

    return ExtractionResult(
        rate_card=rate_card_df,
        accessorial_costs=accessorial_costs_df,
        zones=zones_df,
        address_zones=address_zones_df,
        miscellaneous_pricing_rule=miscellaneous_pricing_rule,
        pricing_rounding_rules=pricing_rounding_rules,
        rate_card_header_row_count=header_row_count,
        source_file=file_path,
        output_file=output_file,
    )


def run_interactive_extraction() -> ExtractionResult:
    input_files = list_input_files()
    if not input_files:
        raise FileNotFoundError(f"No Excel files found in {INPUT_DIR}")

    print("Files available in the input folder:")
    selected_file = prompt_choice(
        "Which file should be processed?",
        [path.name for path in input_files],
        default=input_files[0].name if len(input_files) == 1 else None,
    )
    file_path = INPUT_DIR / selected_file

    sheet_names = list_workbook_sheets(file_path)
    sheet_mapping = {
        "rate_card": prompt_sheet_name(
            sheet_names,
            "Rate Card",
            DEFAULT_SHEETS["rate_card"],
        ),
        "accessorial_costs": prompt_sheet_name(
            sheet_names,
            "Accessorial Costs",
            DEFAULT_SHEETS["accessorial_costs"],
        ),
        "zones": prompt_sheet_name(
            sheet_names,
            "Zones",
            DEFAULT_SHEETS["zones"],
        ),
        "address_zones": prompt_sheet_name(
            sheet_names,
            "Address Zones",
            DEFAULT_SHEETS["address_zones"],
        ),
    }

    result = extract_rate_card_workbook(file_path, sheet_mapping)

    print(f"\nSaved extracted workbook to: {result.output_file}")
    print(f"Rate Card rows: {len(result.rate_card)}")
    print(f"Rate Card column header levels: {result.rate_card_header_row_count}")
    print(f"Accessorial Costs rows: {len(result.accessorial_costs)}")
    print(f"Zones rows: {len(result.zones)}")
    print(f"Address Zones rows: {len(result.address_zones)}")
    print(f"Miscellaneous pricing rule: {result.miscellaneous_pricing_rule}")
    print(
        "Pricing rounding: "
        f"shipment={result.pricing_rounding_rules.price_per_shipment_decimals}, "
        f"100kg={result.pricing_rounding_rules.price_per_100kg_decimals}"
    )

    return result


if __name__ == "__main__":
    extraction_result = run_interactive_extraction()
    miscellaneous_pricing_rule = extraction_result.miscellaneous_pricing_rule
    pricing_rounding_rules = extraction_result.pricing_rounding_rules
