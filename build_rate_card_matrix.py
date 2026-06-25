"""
Transform an extracted Rate Card dataframe into a matrix-format workbook.

Reads from the processing folder (or accepts a dataframe directly) and writes
the matrix rate card to the output folder.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from process_accessorial_costs import write_accessorial_costs_sheet
from process_zones import load_zones_from_processing_file, write_zones_txt

from project_paths import OUTPUT_DIR, PROCESSING_DIR

COST_NAME_ROW = 1
APPLY_IF_ROW = 2
RATE_BY_ROW = 3
WEIGHT_BRACKET_ROW = 4
COLUMN_HEADER_ROW = 5
DATA_START_ROW = 6

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="1F4E78")
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
PRICE_NUMBER_FORMAT = "0.00"

SHIPMENT_BLOCKS = {"Lane / Zoning", "Service"}
IGNORED_COLUMNS = {"Owner", "Valid from", "Valid to", "Rate Card Name"}
EXCLUDED_SHIPMENT_COLUMNS = {"Currency"}

COLUMN_RENAMES = {
    "Origin  Zone": "Origin Postal Code Zone",
    "Origin Zone": "Origin Postal Code Zone",
    "Destination Zone": "Destination Postal Code Zone",
}

PRICING_TYPES = {"price per shipment", "price per 100kg", "max"}


@dataclass
class MatrixCostColumn:
    service_name: str
    source_column: tuple[str, str, str]
    bracket_label: str
    rate_unit: str


@dataclass
class MatrixCostBlock:
    service_name: str
    cost_name: str
    columns: list[MatrixCostColumn] = field(default_factory=list)
    bulkiness_value: str | None = None


@dataclass
class MatrixBuildResult:
    matrix_path: Path
    row_count: int
    shipment_column_count: int
    cost_block_count: int
    zones_path: Path | None = None
    zones_corporate_row_count: int = 0
    zones_address_row_count: int = 0


def _normalize_column_label(label: str) -> str:
    return re.sub(r"\s+", " ", label.replace("\n", " ")).strip()


def _clean_bottom_label(label: str) -> str:
    return re.sub(r"_\d+$", "", _normalize_column_label(label))


def _find_column_by_bottom_label(
    columns: pd.Index,
    *labels: str,
) -> tuple[str, str, str] | None:
    normalized_labels = {_normalize_column_label(label).lower() for label in labels}
    for column in columns:
        bottom = _normalize_column_label(str(column[2])).lower()
        if bottom in normalized_labels:
            return column
    return None


def _format_postal_code_zone(iso_value: object, zone_value: object) -> str:
    iso_text = "" if pd.isna(iso_value) else str(iso_value).strip()
    zone_text = "" if pd.isna(zone_value) else str(zone_value).strip()
    if zone_text.endswith(".0"):
        zone_text = zone_text[:-2]
    if iso_text and zone_text:
        return f"{iso_text} Zone {zone_text}"
    return iso_text or zone_text


def extract_rate_card_zone_names(rate_card: pd.DataFrame) -> set[str]:
    origin_iso_column = _find_column_by_bottom_label(rate_card.columns, "Origin ISO")
    origin_zone_column = _find_column_by_bottom_label(
        rate_card.columns,
        "Origin  Zone",
        "Origin Zone",
    )
    destination_iso_column = _find_column_by_bottom_label(
        rate_card.columns,
        "Destination ISO",
    )
    destination_zone_column = _find_column_by_bottom_label(
        rate_card.columns,
        "Destination Zone",
    )

    zone_names: set[str] = set()
    for _, row in rate_card.iterrows():
        for iso_column, zone_column in (
            (origin_iso_column, origin_zone_column),
            (destination_iso_column, destination_zone_column),
        ):
            if iso_column is None or zone_column is None:
                continue
            zone_name = _format_postal_code_zone(row[iso_column], row[zone_column])
            if zone_name:
                zone_names.add(zone_name)

    return zone_names


def _parse_weight_bracket(column_label: str, pricing_type: str = "") -> list[str]:
    label = _clean_bottom_label(column_label)
    label_lower = label.lower()
    pricing_lower = pricing_type.lower()

    if label_lower == "minimum":
        return ["MIN"]

    if "up to" in label_lower or "max price" in label_lower:
        match = re.search(r"(\d+(?:[.,]\d+)?)\s*kg", label_lower)
        if match:
            upper_bound = int(float(match.group(1).replace(",", ".")))
            if pricing_lower == "max":
                return ["MAX"]
            return [f"<={upper_bound}", "MAX"]

    numbers = re.findall(r"(\d+(?:[.,]\d+)?)", label)
    if numbers:
        upper_bound = int(float(numbers[-1].replace(",", ".")))
        return [f"<={upper_bound}"]

    return [label]


def _get_rate_unit(pricing_type: str) -> str:
    pricing_lower = pricing_type.lower()
    if "price per shipment" in pricing_lower:
        return "Flat"
    match = re.search(r"price per (\d+)\s*kg", pricing_lower)
    if match:
        return f"p/{match.group(1)} units"
    if pricing_lower == "max":
        return "Flat"
    return ""


def _build_shipment_columns(columns: pd.Index) -> list[tuple[str, str, str]]:
    shipment_columns: list[tuple[str, str, str]] = []
    for column in columns:
        block_name = str(column[0])
        bottom_label = _normalize_column_label(str(column[2]))
        if block_name not in SHIPMENT_BLOCKS:
            continue
        if bottom_label in EXCLUDED_SHIPMENT_COLUMNS:
            continue
        shipment_columns.append(column)
    return shipment_columns


def _shipment_output_name(column: tuple[str, str, str]) -> str:
    bottom_label = _normalize_column_label(str(column[2]))
    return COLUMN_RENAMES.get(bottom_label, bottom_label)


def _can_merge_groupage_blocks(first: MatrixCostBlock, second: MatrixCostBlock) -> bool:
    if not first.columns or not second.columns:
        return False

    first_service = first.service_name.lower()
    second_service = second.service_name.lower()
    if "groupage (32000)" not in first_service:
        return False
    if "bundled groupage" not in second_service or "32000" not in second_service:
        return False
    if first.columns[-1].bracket_label != "<=3000":
        return False
    if second.columns[0].bracket_label != "<=4000":
        return False
    return True


def _merge_groupage_blocks(
    first: MatrixCostBlock,
    second: MatrixCostBlock,
) -> MatrixCostBlock:
    merged_service = "Groupage (32000) & Bundled Groupage (32000)"
    merged_columns = [
        MatrixCostColumn(
            service_name=merged_service,
            source_column=column.source_column,
            bracket_label=column.bracket_label,
            rate_unit=column.rate_unit,
        )
        for column in first.columns + second.columns
    ]
    return MatrixCostBlock(
        service_name=merged_service,
        cost_name=f"Transport cost ({merged_service})",
        columns=merged_columns,
    )


def _merge_combinable_cost_blocks(blocks: list[MatrixCostBlock]) -> list[MatrixCostBlock]:
    if not blocks:
        return blocks

    merged_blocks: list[MatrixCostBlock] = []
    index = 0
    while index < len(blocks):
        current_block = blocks[index]
        if index + 1 < len(blocks) and _can_merge_groupage_blocks(
            current_block,
            blocks[index + 1],
        ):
            merged_blocks.append(_merge_groupage_blocks(current_block, blocks[index + 1]))
            index += 2
            continue

        merged_blocks.append(current_block)
        index += 1

    return merged_blocks


def _build_cost_blocks(columns: pd.Index) -> list[MatrixCostBlock]:
    blocks: list[MatrixCostBlock] = []
    current_service: str | None = None
    current_block: MatrixCostBlock | None = None

    for column in columns:
        service_name = str(column[0]).strip()
        pricing_type = str(column[1]).strip()
        pricing_type_normalized = pricing_type.lower()
        bottom_label = _normalize_column_label(str(column[2]))

        if bottom_label in IGNORED_COLUMNS or service_name in SHIPMENT_BLOCKS or not service_name:
            continue
        if pricing_type_normalized not in PRICING_TYPES:
            continue

        if current_service != service_name:
            if current_block is not None:
                blocks.append(current_block)
            current_service = service_name
            current_block = MatrixCostBlock(
                service_name=service_name,
                cost_name=f"Transport cost ({service_name})",
            )

        rate_unit = _get_rate_unit(pricing_type_normalized)
        for bracket_label in _parse_weight_bracket(bottom_label, pricing_type_normalized):
            current_block.columns.append(
                MatrixCostColumn(
                    service_name=service_name,
                    source_column=column,
                    bracket_label=bracket_label,
                    rate_unit=rate_unit,
                )
            )

    if current_block is not None:
        blocks.append(current_block)

    return _merge_combinable_cost_blocks(blocks)


def _normalize_bulkiness(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        numeric = float(text.replace(",", "."))
        if numeric == int(numeric):
            return str(int(numeric))
    except ValueError:
        pass
    return text


def _bulkiness_sort_key(value: str) -> tuple[int, float | str]:
    try:
        return (0, float(value))
    except ValueError:
        return (1, value)


def _is_ltl_block(block: MatrixCostBlock) -> bool:
    return "ltl (33000)" in block.service_name.lower()


def _split_ltl_blocks_by_bulkiness(
    blocks: list[MatrixCostBlock],
    rate_card: pd.DataFrame,
) -> list[MatrixCostBlock]:
    bulkiness_column = _find_column_by_bottom_label(rate_card.columns, "Bulkiness")
    if bulkiness_column is None:
        return blocks

    bulkiness_values = sorted(
        {
            normalized
            for normalized in (
                _normalize_bulkiness(value) for value in rate_card[bulkiness_column].dropna()
            )
            if normalized is not None
        },
        key=_bulkiness_sort_key,
    )

    split_blocks: list[MatrixCostBlock] = []
    for block in blocks:
        if not _is_ltl_block(block):
            split_blocks.append(block)
            continue

        for bulkiness_value in bulkiness_values:
            service_name = f"LTL (33000) Bulkiness {bulkiness_value}"
            split_blocks.append(
                MatrixCostBlock(
                    service_name=service_name,
                    cost_name=f"Transport cost (LTL (33000) Bulkiness {bulkiness_value})",
                    columns=[
                        MatrixCostColumn(
                            service_name=service_name,
                            source_column=column.source_column,
                            bracket_label=column.bracket_label,
                            rate_unit=column.rate_unit,
                        )
                        for column in block.columns
                    ],
                    bulkiness_value=bulkiness_value,
                )
            )

    return split_blocks


def _find_ftl_max_source_column(blocks: list[MatrixCostBlock]) -> tuple[str, str, str] | None:
    for block in blocks:
        if "ftl" not in block.service_name.lower():
            continue
        for column in block.columns:
            if column.bracket_label == "MAX":
                return column.source_column
    return None


def _ltl_bulkiness_blocks_missing_max(blocks: list[MatrixCostBlock]) -> bool:
    return any(
        block.bulkiness_value is not None
        and not any(column.bracket_label == "MAX" for column in block.columns)
        for block in blocks
    )


def _add_ltl_max_from_ftl(
    blocks: list[MatrixCostBlock],
    ftl_max_source_column: tuple[str, str, str] | None,
) -> list[MatrixCostBlock]:
    if ftl_max_source_column is None:
        return blocks

    updated_blocks: list[MatrixCostBlock] = []
    for block in blocks:
        if block.bulkiness_value is None:
            updated_blocks.append(block)
            continue
        if any(column.bracket_label == "MAX" for column in block.columns):
            updated_blocks.append(block)
            continue

        updated_blocks.append(
            MatrixCostBlock(
                service_name=block.service_name,
                cost_name=block.cost_name,
                columns=[
                    *block.columns,
                    MatrixCostColumn(
                        service_name=block.service_name,
                        source_column=ftl_max_source_column,
                        bracket_label="MAX",
                        rate_unit="Flat",
                    ),
                ],
                bulkiness_value=block.bulkiness_value,
            )
        )

    return updated_blocks


def prompt_yes_no(question: str, default: bool = False) -> bool:
    default_hint = "Y/n" if default else "y/N"
    while True:
        raw = input(f"{question} [{default_hint}]: ").strip().lower()
        if not raw:
            return default
        if raw in {"y", "yes"}:
            return True
        if raw in {"n", "no"}:
            return False
        print("Please enter y or n.")


def prompt_add_ltl_max_from_ftl(rate_card: pd.DataFrame) -> bool:
    cost_blocks = _build_cost_blocks(rate_card.columns)
    cost_blocks = _split_ltl_blocks_by_bulkiness(cost_blocks, rate_card)
    if not _ltl_bulkiness_blocks_missing_max(cost_blocks):
        return False
    if _find_ftl_max_source_column(_build_cost_blocks(rate_card.columns)) is None:
        print("LTL has no MAX column and no FTL MAX source was found — skipping LTL MAX.")
        return False

    return prompt_yes_no(
        "Add MAX (Flat) to all LTL bulkiness blocks using values from FTL MAX?",
        default=False,
    )


def _build_cost_columns(
    cost_blocks: list[MatrixCostBlock],
    currency_column: tuple[str, str, str],
) -> list[MatrixCostColumn]:
    cost_columns: list[MatrixCostColumn] = []
    for block in cost_blocks:
        cost_columns.append(
            MatrixCostColumn(
                service_name=block.service_name,
                source_column=currency_column,
                bracket_label="Currency",
                rate_unit="Currency",
            )
        )
        cost_columns.extend(block.columns)
    return cost_columns


def _build_block_lookup(cost_blocks: list[MatrixCostBlock]) -> dict[str, MatrixCostBlock]:
    return {block.service_name: block for block in cost_blocks}


def _row_matches_bulkiness(
    row: pd.Series,
    bulkiness_column: tuple[str, str, str],
    bulkiness_value: str | None,
) -> bool:
    if bulkiness_value is None:
        return True
    return _normalize_bulkiness(row[bulkiness_column]) == bulkiness_value


def _is_empty_value(value: object) -> bool:
    if value is None or pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _cost_block_has_data(
    block: MatrixCostBlock,
    rate_card: pd.DataFrame,
    bulkiness_column: tuple[str, str, str] | None,
) -> bool:
    for _, row in rate_card.iterrows():
        if bulkiness_column is not None and block.bulkiness_value is not None:
            if not _row_matches_bulkiness(row, bulkiness_column, block.bulkiness_value):
                continue

        for column in block.columns:
            if not _is_empty_value(row[column.source_column]):
                return True

    return False


def _filter_empty_cost_blocks(
    blocks: list[MatrixCostBlock],
    rate_card: pd.DataFrame,
) -> list[MatrixCostBlock]:
    bulkiness_column = _find_column_by_bottom_label(rate_card.columns, "Bulkiness")
    return [
        block
        for block in blocks
        if _cost_block_has_data(block, rate_card, bulkiness_column)
    ]


def _build_shipment_values(
    row: pd.Series,
    shipment_columns: list[tuple[str, str, str]],
    origin_iso_column: tuple[str, str, str],
    origin_zone_column: tuple[str, str, str],
    destination_iso_column: tuple[str, str, str],
    destination_zone_column: tuple[str, str, str],
) -> list[object]:
    values: list[object] = []

    for column in shipment_columns:
        output_name = _shipment_output_name(column)
        if output_name == "Origin Postal Code Zone":
            values.append(
                _format_postal_code_zone(
                    row[origin_iso_column],
                    row[origin_zone_column],
                )
            )
            continue
        if output_name == "Destination Postal Code Zone":
            values.append(
                _format_postal_code_zone(
                    row[destination_iso_column],
                    row[destination_zone_column],
                )
            )
            continue
        values.append(row[column])

    return values


def _build_matrix_rows(
    rate_card: pd.DataFrame,
    shipment_columns: list[tuple[str, str, str]],
    cost_blocks: list[MatrixCostBlock],
) -> tuple[list[str], list[list[object]]]:
    origin_iso_column = _find_column_by_bottom_label(rate_card.columns, "Origin ISO")
    origin_zone_column = _find_column_by_bottom_label(
        rate_card.columns,
        "Origin  Zone",
        "Origin Zone",
    )
    destination_iso_column = _find_column_by_bottom_label(
        rate_card.columns,
        "Destination ISO",
    )
    destination_zone_column = _find_column_by_bottom_label(
        rate_card.columns,
        "Destination Zone",
    )
    currency_column = _find_column_by_bottom_label(rate_card.columns, "Currency")

    if currency_column is None:
        raise ValueError("Currency column not found in Rate Card dataframe.")

    bulkiness_column = _find_column_by_bottom_label(rate_card.columns, "Bulkiness")
    block_lookup = _build_block_lookup(cost_blocks)
    shipment_headers = [_shipment_output_name(column) for column in shipment_columns]
    cost_columns = _build_cost_columns(cost_blocks, currency_column)

    header_rows: list[list[object]] = []
    for header_index in range(5):
        row_values: list[object] = []
        if header_index == 4:
            row_values.extend(shipment_headers)
        else:
            row_values.extend([""] * len(shipment_headers))

        for column in cost_columns:
            if header_index == 0:
                block = next(
                    block for block in cost_blocks if block.service_name == column.service_name
                )
                row_values.append(block.cost_name)
            elif header_index in (1, 2):
                row_values.append("")
            elif header_index == 3:
                row_values.append("" if column.bracket_label == "Currency" else column.bracket_label)
            elif header_index == 4:
                row_values.append(column.rate_unit)
            else:
                row_values.append("")

        header_rows.append(row_values)

    data_rows: list[list[object]] = []
    for _, row in rate_card.iterrows():
        data_row = _build_shipment_values(
            row,
            shipment_columns,
            origin_iso_column,
            origin_zone_column,
            destination_iso_column,
            destination_zone_column,
        )

        for column in cost_columns:
            block = block_lookup[column.service_name]
            if bulkiness_column is not None and not _row_matches_bulkiness(
                row,
                bulkiness_column,
                block.bulkiness_value,
            ):
                data_row.append(None)
                continue

            if column.bracket_label == "Currency":
                data_row.append(row[currency_column])
                continue

            data_row.append(row[column.source_column])

        data_rows.append(data_row)

    return shipment_headers, header_rows + data_rows


def _merge_cost_name_cells(worksheet, shipment_column_count: int, cost_blocks: list[MatrixCostBlock]) -> None:
    current_column = shipment_column_count + 1
    for block in cost_blocks:
        block_width = 1 + len(block.columns)
        if block_width > 1:
            worksheet.merge_cells(
                start_row=COST_NAME_ROW,
                start_column=current_column,
                end_row=COST_NAME_ROW,
                end_column=current_column + block_width - 1,
            )
        current_column += block_width


BOLD_SHIPMENT_HEADERS = {"Origin Postal Code Zone", "Destination Postal Code Zone"}


def _apply_worksheet_formatting(
    worksheet,
    shipment_column_count: int,
    shipment_headers: list[str],
    cost_columns: list[MatrixCostColumn],
    total_rows: int,
) -> None:
    total_columns = shipment_column_count + len(cost_columns)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_index in range(1, total_rows + 1):
        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = THIN_BORDER

            if row_index == COST_NAME_ROW and column_index > shipment_column_count:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = center
            elif row_index == WEIGHT_BRACKET_ROW and column_index > shipment_column_count:
                cell.fill = SUBHEADER_FILL
                cell.font = SUBHEADER_FONT
                cell.alignment = center
            elif row_index == COLUMN_HEADER_ROW and column_index <= shipment_column_count:
                header_name = shipment_headers[column_index - 1]
                cell.alignment = left
                if header_name in BOLD_SHIPMENT_HEADERS:
                    cell.font = BOLD_FONT
            elif row_index >= DATA_START_ROW and column_index > shipment_column_count:
                cost_column = cost_columns[column_index - shipment_column_count - 1]
                if cost_column.bracket_label != "Currency":
                    cell.number_format = PRICE_NUMBER_FORMAT
                    cell.alignment = center

    worksheet.sheet_view.showGridLines = True


def build_rate_card_matrix(
    rate_card: pd.DataFrame,
    output_path: Path,
    accessorial_costs: pd.DataFrame | None = None,
    add_ltl_max_from_ftl: bool = False,
) -> MatrixBuildResult:
    if not isinstance(rate_card.columns, pd.MultiIndex):
        raise ValueError("Rate Card dataframe must have a 3-level MultiIndex header.")

    shipment_columns = _build_shipment_columns(rate_card.columns)
    cost_blocks = _build_cost_blocks(rate_card.columns)
    ftl_max_source_column = _find_ftl_max_source_column(cost_blocks)
    cost_blocks = _split_ltl_blocks_by_bulkiness(cost_blocks, rate_card)
    if add_ltl_max_from_ftl:
        cost_blocks = _add_ltl_max_from_ftl(cost_blocks, ftl_max_source_column)
    cost_blocks = _filter_empty_cost_blocks(cost_blocks, rate_card)
    shipment_headers, matrix_rows = _build_matrix_rows(rate_card, shipment_columns, cost_blocks)

    currency_column = _find_column_by_bottom_label(rate_card.columns, "Currency")
    cost_columns = _build_cost_columns(cost_blocks, currency_column)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Rate Card"

    for row_index, row_values in enumerate(matrix_rows, start=1):
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if value == "":
                cell.value = None

    _merge_cost_name_cells(worksheet, len(shipment_columns), cost_blocks)
    _apply_worksheet_formatting(
        worksheet,
        len(shipment_columns),
        shipment_headers,
        cost_columns,
        len(matrix_rows),
    )

    if accessorial_costs is not None:
        write_accessorial_costs_sheet(workbook, accessorial_costs)

    workbook.save(output_path)

    return MatrixBuildResult(
        matrix_path=output_path,
        row_count=len(rate_card),
        shipment_column_count=len(shipment_columns),
        cost_block_count=len(cost_blocks),
    )


def load_extracted_rate_card(processing_file: Path) -> pd.DataFrame:
    return pd.read_excel(processing_file, sheet_name="Rate Card", header=[0, 1, 2])


def load_extracted_accessorial_costs(processing_file: Path) -> pd.DataFrame:
    return pd.read_excel(processing_file, sheet_name="Accessorial Costs", header=0)


def build_matrix_from_extraction_result(
    extraction_result,
    output_file: Path | None = None,
    add_ltl_max_from_ftl: bool = False,
) -> MatrixBuildResult:
    if output_file is None:
        output_file = OUTPUT_DIR / f"{extraction_result.source_file.stem}_matrix.xlsx"
    matrix_result = build_rate_card_matrix(
        extraction_result.rate_card,
        output_file,
        accessorial_costs=extraction_result.accessorial_costs,
        add_ltl_max_from_ftl=add_ltl_max_from_ftl,
    )

    zones_output = OUTPUT_DIR / f"{extraction_result.source_file.stem}_zones.txt"
    used_zone_names = extract_rate_card_zone_names(extraction_result.rate_card)
    _, corporate_count, address_count = write_zones_txt(
        extraction_result.zones,
        extraction_result.address_zones,
        zones_output,
        used_zone_names=used_zone_names,
    )
    return MatrixBuildResult(
        matrix_path=matrix_result.matrix_path,
        row_count=matrix_result.row_count,
        shipment_column_count=matrix_result.shipment_column_count,
        cost_block_count=matrix_result.cost_block_count,
        zones_path=zones_output,
        zones_corporate_row_count=corporate_count,
        zones_address_row_count=address_count,
    )


def build_matrix_from_processing_file(
    processing_file: Path,
    output_file: Path | None = None,
    add_ltl_max_from_ftl: bool = False,
) -> MatrixBuildResult:
    rate_card = load_extracted_rate_card(processing_file)
    accessorial_costs = load_extracted_accessorial_costs(processing_file)
    if output_file is None:
        output_file = OUTPUT_DIR / f"{processing_file.stem}_matrix.xlsx"
    matrix_result = build_rate_card_matrix(
        rate_card,
        output_file,
        accessorial_costs=accessorial_costs,
        add_ltl_max_from_ftl=add_ltl_max_from_ftl,
    )

    zones_output = OUTPUT_DIR / f"{processing_file.stem}_zones.txt"
    zones_df, address_zones_df = load_zones_from_processing_file(processing_file)
    used_zone_names = extract_rate_card_zone_names(rate_card)
    _, corporate_count, address_count = write_zones_txt(
        zones_df,
        address_zones_df,
        zones_output,
        used_zone_names=used_zone_names,
    )
    return MatrixBuildResult(
        matrix_path=matrix_result.matrix_path,
        row_count=matrix_result.row_count,
        shipment_column_count=matrix_result.shipment_column_count,
        cost_block_count=matrix_result.cost_block_count,
        zones_path=zones_output,
        zones_corporate_row_count=corporate_count,
        zones_address_row_count=address_count,
    )


def list_processing_files() -> list[Path]:
    if not PROCESSING_DIR.exists():
        return []
    return sorted(PROCESSING_DIR.glob("*_extracted.xlsx"))


def run_interactive_matrix_build() -> MatrixBuildResult:
    processing_files = list_processing_files()
    if not processing_files:
        raise FileNotFoundError(f"No extracted files found in {PROCESSING_DIR}")

    print("Extracted files in processing folder:")
    for index, file_path in enumerate(processing_files, start=1):
        print(f"  {index}. {file_path.name}")

    while True:
        raw = input("Enter file number to convert: ").strip()
        if raw.isdigit():
            choice_index = int(raw) - 1
            if 0 <= choice_index < len(processing_files):
                selected_file = processing_files[choice_index]
                break
        print("Invalid choice. Try again.")

    rate_card = load_extracted_rate_card(selected_file)
    add_ltl_max_from_ftl = prompt_add_ltl_max_from_ftl(rate_card)
    result = build_matrix_from_processing_file(
        selected_file,
        add_ltl_max_from_ftl=add_ltl_max_from_ftl,
    )
    print(f"\nSaved matrix rate card to: {result.matrix_path}")
    print(f"Data rows: {result.row_count}")
    print(f"Shipment columns: {result.shipment_column_count}")
    print(f"Cost blocks: {result.cost_block_count}")
    if result.zones_path is not None:
        print(f"Saved zones file to: {result.zones_path}")
        print(f"Corporate zone rows: {result.zones_corporate_row_count}")
        print(f"Address zone rows: {result.zones_address_row_count}")
    return result


if __name__ == "__main__":
    build_result = run_interactive_matrix_build()
