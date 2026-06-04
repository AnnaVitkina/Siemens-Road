"""
Transform accessorial costs into the Rate Agreement format and write to workbook.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ACCESSORIAL_SHEET_NAME = "Accessorial Costs"

OUTPUT_COLUMNS = [
    "Rate Card cost name",
    "Rate Agreement cost name",
    "Price",
    "Currency",
    "Apply if",
    "Rate by",
]

ACCESSORIAL_COST_MAPPINGS: dict[str, str] = {
    "2nd driver for urgent shipments": "Second driver (2nd driver for urgent shipments)",
    "2nd delivery caused by Siemens": "Second Delivery (2nd delivery caused by Siemens)",
    "Skeleton Box / Gitterbox Exchange in Germany": (
        "Skeleton Box Exchange Fee (Skeleton Box / Gitterbox Exchange in Germany)"
    ),
    "Demurrage after 2 hrs max 8 hrs": "Demurrage Fee (Demurrage after 2 hrs max 8 hrs)",
    "Demurrage after 8 hrs": "Demurrage after 8 hours",
    "LTL - truck interior overheight up to 265 cm": (
        "Overheight Fee (LTL - truck interior overheight; FTL - truck interior overheight)"
    ),
    "LTL - truck interior overheight up to 300 cm": (
        "Overheight Fee (LTL - truck interior overheight; FTL - truck interior overheight)"
    ),
    "LTL - delivery on the same day - standard truck interior height 245 cm": (
        "Same Day Delivery (LTL - delivery on the same day; FTL - delivery on the same day)"
    ),
    "LTL - delivery on the same day - truck interior overheight up to 265 cm": (
        "Same Day Delivery (LTL - delivery on the same day; FTL - delivery on the same day)"
    ),
    "LTL - delivery on the same day - truck interior overheight up to 300 cm": (
        "Same Day Delivery (LTL - delivery on the same day; FTL - delivery on the same day)"
    ),
    "LTL - delivery on fixed date - standard truck interior height 245 cm": (
        "Fixed Day Delivery Fee (LTL - delivery on fixed date; FTL - delivery on fixed date)"
    ),
    "LTL - delivery on fixed date - truck interior overheight up to 265 cm": (
        "Fixed Day Delivery Fee (LTL - delivery on fixed date; FTL - delivery on fixed date)"
    ),
    "LTL - delivery on fixed date - truck interior overheight up to 300 cm": (
        "Fixed Day Delivery Fee (LTL - delivery on fixed date; FTL - delivery on fixed date)"
    ),
    "LTL - tail lift - standard truck interior height 245 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "LTL - tail lift - truck interior overheight up to 265 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "LTL - tail lift - delivery on the same day - standard truck interior height 245 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "LTL - tail lift - delivery on the same day - truck interior overheight up to 265 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "LTL - tail lift - delivery on fixed date - standard truck interior height 245 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "LTL - tail lift - delivery on fixed date - truck interior overheight up to 265 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "LTL - tail lift - delivery on fixed date - truck interior overheight up to 300 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "FTL - truck interior overheight up to 265 cm": (
        "Overheight Fee (LTL - truck interior overheight; FTL - truck interior overheight)"
    ),
    "FTL - truck interior overheight up to 300 cm": (
        "Overheight Fee (LTL - truck interior overheight; FTL - truck interior overheight)"
    ),
    "FTL - delivery on the same day - standard truck interior height 245 cm": (
        "Same Day Delivery (LTL - delivery on the same day; FTL - delivery on the same day)"
    ),
    "FTL - delivery on the same day - truck interior overheight up to 265 cm": (
        "Same Day Delivery (LTL - delivery on the same day; FTL - delivery on the same day)"
    ),
    "FTL - delivery on the same day - truck interior overheight up to 300 cm": (
        "Same Day Delivery (LTL - delivery on the same day; FTL - delivery on the same day)"
    ),
    "FTL - delivery on fixed date - standard truck interior height 245 cm": (
        "Fixed Day Delivery Fee (LTL - delivery on fixed date; FTL - delivery on fixed date)"
    ),
    "FTL - delivery on fixed date - truck interior overheight up to 265 cm": (
        "Fixed Day Delivery Fee (LTL - delivery on fixed date; FTL - delivery on fixed date)"
    ),
    "FTL - delivery on fixed date - truck interior overheight up to 300 cm": (
        "Fixed Day Delivery Fee (LTL - delivery on fixed date; FTL - delivery on fixed date)"
    ),
    "FTL - tail lift - standard truck interior height 245 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "FTL - tail lift - truck interior overheight up to 265 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "FTL - tail lift - delivery on the same day - standard truck interior height 245 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "FTL - tail lift - delivery on the same day - truck interior overheight up to 265 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "FTL - tail lift - delivery on fixed date - standard truck interior height 245 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "FTL - tail lift - delivery on fixed date - truck interior overheight up to 265 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "FTL - tail lift - delivery on fixed date - truck interior overheight up to 300 cm": (
        "Tail Lift Fee (LTL - tail lift; FTL - tail lift)"
    ),
    "Transit declaration: T1 - Document- Handling": (
        "T1 Document Fee (Transit declaration: T1 - Document- Handling)"
    ),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
PRICE_NUMBER_FORMAT = "#,##0.00"


def _normalize_service_name(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def _format_service_code(value: object) -> str | None:
    if pd.isna(value):
        return None
    numeric = float(value)
    if numeric == int(numeric):
        return str(int(numeric))
    return str(value).strip()


def _build_apply_if(service_code: object) -> str:
    formatted_code = _format_service_code(service_code)
    if formatted_code is None:
        return ""
    return f"Service_Code equals {formatted_code}"


def _normalize_price(value: object) -> object:
    if pd.isna(value):
        return value
    if isinstance(value, str):
        return value.strip()
    try:
        numeric = float(value)
        if numeric == int(numeric):
            return int(numeric)
        return numeric
    except (TypeError, ValueError):
        return value


def _combine_apply_if(apply_if_values: list[str]) -> str:
    parts = [value for value in apply_if_values if value]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    return " or ".join(parts)


def _combine_rate_card_names(names: list[str]) -> str:
    return " | ".join(names)


def _consolidate_accessorial_costs(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    groups: dict[tuple[object, ...], list[int]] = {}
    group_order: list[tuple[object, ...]] = []

    for row_index, row in df.iterrows():
        group_key = (
            row["Rate Agreement cost name"],
            _normalize_price(row["Price"]),
            row["Currency"],
            row["Rate by"],
        )
        if group_key not in groups:
            groups[group_key] = []
            group_order.append(group_key)
        groups[group_key].append(row_index)

    records: list[dict[str, object]] = []
    for group_key in group_order:
        group_rows = df.loc[groups[group_key]]
        records.append(
            {
                "Rate Card cost name": _combine_rate_card_names(
                    group_rows["Rate Card cost name"].tolist()
                ),
                "Rate Agreement cost name": group_key[0],
                "Price": group_rows.iloc[0]["Price"],
                "Currency": group_key[2],
                "Apply if": _combine_apply_if(group_rows["Apply if"].tolist()),
                "Rate by": group_key[3],
            }
        )

    return pd.DataFrame(records, columns=OUTPUT_COLUMNS)


def process_accessorial_costs(accessorial_df: pd.DataFrame) -> pd.DataFrame:
    source_by_name: dict[str, pd.Series] = {}
    for _, row in accessorial_df.iterrows():
        service_name = _normalize_service_name(row["Service_Name"])
        source_by_name[service_name] = row

    mapped_records: list[dict[str, object]] = []
    for rate_card_name, agreement_name in ACCESSORIAL_COST_MAPPINGS.items():
        row = source_by_name.get(rate_card_name)
        if row is None:
            continue

        mapped_records.append(
            {
                "Rate Card cost name": rate_card_name,
                "Rate Agreement cost name": agreement_name,
                "Price": row["Price"],
                "Currency": row["Currency"],
                "Apply if": _build_apply_if(row["Service_Code"]),
                "Rate by": row["Cost_Unit"],
            }
        )

    unmapped_records: list[dict[str, object]] = []
    for service_name, row in source_by_name.items():
        if service_name in ACCESSORIAL_COST_MAPPINGS:
            continue

        unmapped_records.append(
            {
                "Rate Card cost name": service_name,
                "Rate Agreement cost name": "",
                "Price": row["Price"],
                "Currency": row["Currency"],
                "Apply if": _build_apply_if(row["Service_Code"]),
                "Rate by": row["Cost_Unit"],
            }
        )

    consolidated_mapped = _consolidate_accessorial_costs(
        pd.DataFrame(mapped_records, columns=OUTPUT_COLUMNS)
    )
    unmapped = pd.DataFrame(unmapped_records, columns=OUTPUT_COLUMNS)

    if consolidated_mapped.empty:
        return unmapped
    if unmapped.empty:
        return consolidated_mapped

    return pd.concat([consolidated_mapped, unmapped], ignore_index=True)


def _apply_accessorial_formatting(worksheet, row_count: int) -> None:
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    column_widths = {
        1: 55,
        2: 70,
        3: 12,
        4: 10,
        5: 55,
        6: 18,
    }
    for column_index, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for column_index in range(1, len(OUTPUT_COLUMNS) + 1):
        header_cell = worksheet.cell(row=1, column=column_index)
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
        header_cell.alignment = center
        header_cell.border = THIN_BORDER

    for row_index in range(2, row_count + 1):
        for column_index in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = THIN_BORDER
            cell.alignment = left if column_index in (1, 2, 5) else center

            if column_index == 3:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = PRICE_NUMBER_FORMAT

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True


def write_accessorial_costs_sheet(workbook, accessorial_df: pd.DataFrame) -> pd.DataFrame:
    processed = process_accessorial_costs(accessorial_df)
    worksheet = workbook.create_sheet(title=ACCESSORIAL_SHEET_NAME)

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    for row_index, row in enumerate(processed.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    _apply_accessorial_formatting(worksheet, len(processed) + 1)
    return processed
